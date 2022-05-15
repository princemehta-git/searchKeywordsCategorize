# import re
from urllib.request import urlopen, Request
from googlesearch import search
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter import ttk
from PIL import ImageTk, Image
import random
from os import listdir
import pyttsx3
import subprocess


main_window = tk.Tk()

# main_window.geometry("800x500")
main_window.title('Idiot Scraper')
main_window.iconbitmap('icon.ico')
main_window.geometry("900x700")
# random_img = random.choice(listdir('pics'))
# bg = ImageTk.PhotoImage(Image.open('pics\\'+ random_img))
# tk.Label( main_window, image = bg).place(x = 0,y = 0)

def resize_image(event):
    new_width = event.width
    new_height = event.height
    image = copy_of_image.resize((new_width, new_height))
    photo = ImageTk.PhotoImage(image)
    label.config(image = photo)
    label.image = photo #avoid garbage collection

random_img = random.choice(listdir('pics'))
# bg = ImageTk.PhotoImage(Image.open('pics\\'+ random_img))
image = Image.open('pics\\'+ random_img)
# main_window.geometry(str(image.width)+'x'+str(image.height))
# main_window.geometry("900x700")
copy_of_image = image.copy()
photo = ImageTk.PhotoImage(image)
label = ttk.Label(main_window, image = photo)
label.bind('<Configure>', resize_image)
label.pack(fill=tk.BOTH, expand = tk.YES)


file_path_var = tk.StringVar()
url_var = tk.IntVar()

file_path_label = tk.Label(main_window, text='file Path', font=('Bradley Hand ITC', 20, 'bold'), bg = "#000000", fg = '#D1FF00')
file_path_entry = tk.Entry(main_window, textvariable=file_path_var, font=('Calibre', 10, 'italic'), width = 45)

def browse_file():
    filename = askopenfilename(filetypes=(("excel | xlsx file", "*.xlsx"),("All files", " *.* "),))
    file_path_entry.delete(0, 'end')
    file_path_entry.insert(index=0,string=filename)  # add this
    return None

file_path_label.place(relx=0.01, rely=0.05, relwidth=0.2, relheight=0.035)
file_path_entry.place(relx=0.2, rely=0.05, relwidth=0.68, relheight=0.035)
tk.Button(main_window, text="Browse", font=('Lucida Calligraphy', 10, 'bold'), command=browse_file, bg = "#000000", fg = '#FF00D8').place(relx=0.9, rely=0.05, relwidth=0.08, relheight=0.035)





url_label = tk.Label(main_window, text='Till which url:', font=('Bradley Hand ITC', 16, 'bold'), bg = "#000000", fg = '#D1FF00')
url_label.place(relx=0.4, rely=0.3, relheight=0.04, relwidth=0.15)


which_url = ttk.Combobox(main_window, textvariable = url_var)
which_url.place(relx=0.55, rely=0.3, relheight=0.04, relwidth=0.06)

which_url['values'] = [1,2,3,4,5,6,7,8,9,10]

which_url.set(3)





def scrape(url):
    try:
        page = urlopen(url)
        # start_scrape(page)
    except:
        hdr = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36'}
        # hdr = {"User-Agent": "My Agent"}
        req = Request(url, headers=hdr)
        page = urlopen(req)

    scrape = BeautifulSoup(page, 'html.parser')
    scrape_txt = scrape.get_text().lower().strip().strip('.-:;,\'"\)\(|[]{}/\\@!#$%^&*+=*<>?')
    scrape_lst = scrape_txt.lower().split(' ')
    scrape_set = []
    for word in scrape_lst:
        scrape_set.append(word.strip().strip('.-:;,\'"\)\(|[]{}/\\@!#$%^&*+=*<>?').strip(' '))

    scrape_set = set(scrape_set)

    # if 'neet' in scrape_set or 'x science' in scrape_txt or 'x sci' in scrape_txt or 'xi science' in scrape_txt or 'xi sci' in scrape_txt or 'xii science' in scrape_txt or 'xii sci' in scrape_txt or 'jee' in scrape_set or 'iit' in scrape_set or 'medical' in scrape_set:
    #     return('JEE')
    #
    # elif 'law' in scrape_set or 'ias' in scrape_set or 'banking' in scrape_set or 'competitive' in scrape_set or 'upsc' in scrape_set or 'jpsc' in scrape_set or 'ssc' in scrape_set or 'railway' in scrape_set or 'rrb' in scrape_set or 'army' in scrape_set:
    #     return('GOVT')
    #
    # # elif 'neet' in scrape_set or 'x science' in scrape_txt or 'x sci' in scrape_txt or 'xi science' in scrape_txt or 'xi sci' in scrape_txt or 'xii science' in scrape_txt or 'xii sci' in scrape_txt or 'jee' in scrape_set or 'iit' in scrape_set or 'medical' in scrape_set:
    # #     return('JEE')
    # elif 'school' in scrape_set or 'vi to x' in scrape_txt or 'vi to x class' in scrape_txt or 'college' in scrape_set:
    #     return('OTHERS')
    # elif 'hospital' in scrape_set or 'restaurant' in scrape_set or 'hotel' in scrape_set or 'mall' in scrape_set or 'shop' in scrape_set or 'gym' in scrape_set:
    #     return('WRONG')
    # else:
    #     return('NONE')

    with open('jee.txt', 'r', encoding="utf-8") as file:
        jee_list = list(file.readline().lower().split(','))

    with open('govt.txt', 'r', encoding="utf-8") as file:
        govt_list = list(file.readline().lower().split(','))

    with open('others.txt', 'r', encoding="utf-8") as file:
        others_list = list(file.readline().lower().split(','))

    with open('wrong.txt', 'r', encoding="utf-8") as file:
        wrong_list = list(file.readline().lower().split(','))

    for word in jee_list:
        word = word.strip().lower()
        if ' ' in word:
            in_check = scrape_txt
        else:
            in_check = scrape_set

        if word in in_check:
            return ('JEE', word)

    for word in govt_list:
        word = word.strip().lower()
        if ' ' in word:
            in_check = scrape_txt
        else:
            in_check = scrape_set

        if word in in_check:
            return ('GOVT', word)

    for word in others_list:
        word = word.strip().lower()
        if ' ' in word:
            in_check = scrape_txt
        else:
            in_check = scrape_set

        if word in in_check:
            return ('OTHERS', word)

    for word in wrong_list:
        word = word.strip().lower()
        if ' ' in word:
            in_check = scrape_txt
        else:
            in_check = scrape_set

        if word in in_check:
            return ('WRONG', word)

    return ('NONE', 'NONE')


# print(link)

# def selenium_scrape(link):
#     from selenium import webdriver
#     from webdriver_manager.chrome import ChromeDriverManager
#     options = webdriver.ChromeOptions()
#     # options.headless = True
#     user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.109 Safari/537.36 OPR/84.0.4316.42"
#     options.add_argument(f'user-agent={user_agent}')
#     options.add_argument("--window-size=1920,1080")
#     options.add_argument('--ignore-certificate-errors')
#     options.add_argument('--allow-running-insecure-content')
#     options.add_argument("--disable-extensions")
#     options.add_argument("--proxy-server='direct://'")
#     options.add_argument("--proxy-bypass-list=*")
#     options.add_argument("--start-maximized")
#     options.add_argument('--disable-gpu')
#     options.add_argument('--disable-dev-shm-usage')
#     options.add_argument('--no-sandbox')
#     options.add_argument("--app=https://www.google.com")
#     options.add_experimental_option("useAutomationExtension", False)
#     options.add_experimental_option("excludeSwitches", ["enable-automation"])
#     # options.headless = not bool(is_showBrowser_var.get())
#
#     try:
#         driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
#         driver.implicitly_wait(60)
#         driver.maximize_window()
#     except:
#         # popup('Update Chrome', r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe', 'error')
#         return
#     driver.get(link)
#     print(driver.current_url)
#     page = driver.page_source
#     print(driver.current_url)
#     scrape = BeautifulSoup(page, 'html.parser')
#     scrape_txt = scrape.get_text().lower().strip().strip('.-:;,\'"\)\(|[]{}/\\@!#$%^&*+=*<>?')
#     scrape_lst = scrape_txt.lower().split(' ')
#     scrape_set = []
#     for word in scrape_lst:
#         scrape_set.append(word.strip().strip('.-:;,\'"\)\(|[]{}/\\@!#$%^&*+=*<>?').strip(' '))
#
#     scrape_set = set(scrape_set)
#
#     if 'law' in scrape_set or 'ias' in scrape_set or 'banking' in scrape_set or 'competitive' in scrape_set or 'upsc' in scrape_set or 'jpsc' in scrape_set or 'ssc' in scrape_set or 'railway' in scrape_set or 'rrb' in scrape_set or 'army' in scrape_set:
#         print('govt')
#
#     elif 'neet' in scrape_set or 'x science' in scrape_txt or 'x sci' in scrape_txt or 'xi science' in scrape_txt or 'xi sci' in scrape_txt or 'xii science' in scrape_txt or 'xii sci' in scrape_txt or 'jee' in scrape_set or 'iit' in scrape_set or 'medical' in scrape_set:
#         print('JEE')
#     elif 'school' in scrape_set or 'vi to x' in scrape_txt or 'vi to x class' in scrape_txt or 'college' in scrape_set:
#         print('OTHERS')
#     elif 'hospital' in scrape_set or 'restaurant' in scrape_set or 'hotel' in scrape_set or 'mall' in scrape_set or 'shop' in scrape_set:
#         print('WRONG')
#     else:
#         print('couldnt find any word')


def main():
    wb = load_workbook(file_path_var.get())
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row+1):
        print(row-1)
        if sheet.cell(row = row, column=2).value != None:
            search_word = sheet.cell(row=row, column=2).value.strip() + ' '+ sheet.cell(row=row, column=1).value.strip()
            for trial in range(1,int(which_url.get()) + 1):
                try:
                    link = search(search_word, tld="co.in", num=1, start=int(trial) - 1, stop=1, pause=1.5,
                      lang='english').__next__()
                except:
                    try:
                        link = search(search_word, tld="co.in", num=1, start=int(trial) - 1, stop=1, pause=1.5,
                                      lang='english').__next__()
                    except:
                        continue
                try:
                    output = scrape(link)
                except:
                    output = ('NONE', 'NONE')
                if 'none' not in output[0].lower():
                    sheet.cell(row=row, column=3).value = output[0]
                    sheet.cell(row=row, column= 4 ).value = link
                    sheet.cell(row=row, column=5).value = output[1]
                    wb.save('report.xlsx')
                    break
                else:
                    pass
        # else:
        #     return
    popup('Task Compeleted!!', 'report.xlsx','notification')


def popup(txt, path, error_or_notification):
    global bg1, bg2
    top = tk.Toplevel(main_window)

    if 'err' in error_or_notification.lower():
        top.title('ERROR')
        top.iconbitmap('error.ico')
        bg1 = ImageTk.PhotoImage(Image.open('error.jpg'))
        tk.Label(top, image=bg1).place(x=-2, y=-2)
        # winsound.MessageBeep(winsound.MB_ICONHAND)


    else:
        top.title('Notifier')
        top.iconbitmap('icon.ico')
        bg2 = ImageTk.PhotoImage(Image.open('notification.jpg'))
        tk.Label(top, image=bg2).place(x=-2, y=-2)
        # winsound.MessageBeep(winsound.MB_OK)
    top.geometry("400x200")


    # Create an Entry Widget in the Toplevel window
    tk.Label(top, text=txt, font=('Bradley Hand ITC', 20, 'bold'), bg = "#000000", fg = '#D1FF00').pack()

    # btn for open folder
    tk.Button(top, text = 'Open',font=('Lucida Calligraphy', 10, 'bold'),bg = "#000000", fg = '#FF00D8', command=lambda:subprocess.call('explorer %s'%path, shell=True)).pack(pady=5, side=tk.TOP)


    # Create a Button Widget in the Toplevel Window
    button = tk.Button(top, text="Cancel",font=('Lucida Calligraphy', 10, 'bold'),bg = "#000000", fg = '#FF00D8', command=lambda :top.destroy())
    button.pack(pady=5, side=tk.TOP)
    try:

        engine = pyttsx3.init()
        voices = engine.getProperty('voices')
        engine.setProperty('voice', voices[1].id)
        engine.setProperty("rate", 230)
        engine.say(txt)
        engine.runAndWait()
    except:
        try:
            engine = pyttsx3.init()
            voices = engine.getProperty('voices')
            engine.setProperty('voice', voices[1].id)
            engine.setProperty("rate", 150)
            engine.say(txt)
            engine.runAndWait()
        except:
            pass







start_btn = tk.Button(main_window, text='Start',font=('Lucida Calligraphy', 10, 'bold'),bg = "#000000", fg = '#FF00D8', command=main)
start_btn.place(relx=0.45, rely=0.5, relwidth=0.09, relheight=0.05)

setting_label = tk.Label(main_window, text='Settings:', font=('Bradley Hand ITC', 16, 'bold'), bg = "#000000", fg = '#D1FF00')
setting_label.place(relx=0.01, rely=0.7, relheight=0.04, relwidth=0.15)


tk.Button(main_window, text = 'JEE',font=('Lucida Calligraphy', 10, 'bold'),bg = "#000000", fg = '#FF00D8', command=lambda:subprocess.call('explorer jee.txt', shell=True)).place(relx=0.01, rely=0.8, relwidth=0.15, relheight=0.05)
tk.Button(main_window, text = 'GOVT',font=('Lucida Calligraphy', 10, 'bold'),bg = "#000000", fg = '#FF00D8', command=lambda:subprocess.call('explorer govt.txt', shell=True)).place(relx=0.25, rely=0.8, relwidth=0.15, relheight=0.05)
tk.Button(main_window, text = 'OTHERS',font=('Lucida Calligraphy', 10, 'bold'),bg = "#000000", fg = '#FF00D8', command=lambda:subprocess.call('explorer others.txt', shell=True)).place(relx=0.5, rely=0.8, relwidth=0.15, relheight=0.05)
tk.Button(main_window, text = 'WRONG',font=('Lucida Calligraphy', 10, 'bold'),bg = "#000000", fg = '#FF00D8', command=lambda:subprocess.call('explorer wrong.txt', shell=True)).place(relx=0.8, rely=0.8, relwidth=0.15, relheight=0.05)










main_window.mainloop()